from openpyxl import Workbook
from openpyxl import load_workbook
import time
from openpyxl.chart import BarChart, Series, Reference
from pylab import *
from matplotlib.ticker import MultipleLocator
from matplotlib.ticker import FormatStrFormatter
import os
import WanoCode
import datetime
import  re
def Tiltle2index(title):
    return listTitle.index(title)

def readWANOcode():
    pass


columsPurpose=['SUBJECT']

rb = load_workbook("dyw time.xlsx")
rs = rb['Sheet1']

wanoModel=WanoCode.WanoCode('编码.xlsx')
rs_rows_len = len(list(rs.rows))
rs_columns_len = len(list(rs.columns))

filepath='D:\Anaconda\workspace\ef\\大亚湾'

rb_p=load_workbook('大亚湾电厂信息.xlsx')
rs_p=rb_p['sheet1']



# title list
listTitle=[]
Titlerow=list(list(rs.rows)[0])

for i in range(0,len(Titlerow)):
    listTitle.append(Titlerow[i].value)

list_RULE=['4.1.1','4.1.2','4.1.3','4.1.4','4.1.5','4.1.6','4.1.7','4.1.7.1','4.1.7.2','4.1.7.3','4.1.7.4','4.1.8','4.1.9','other']



# plant list
PlantInfo={}
list_Plant_r=[]
list_Plant=[]
for row in list(rs_p.rows):
    list_Plant.append(row[0].value)
    PlantInfo.update({row[0].value:[row[1].value,row[2].value,row[3].value,row[4].value]})
list_Plant.append('ALLPlant')

dict_wb={}
dict_whole_static={}
dict_trash_wb={}



dict_unit_year={}

def caculate_unityear():
    pass
    for i in range(1991,2018):
        dict_unit_year.update({str(i):0})
    for plant in PlantInfo.items():
        fuelday=datetime.datetime.strptime(plant[1][3], "%Y-%m-%d")
        firstyear=fuelday.year
        days_firstyear=(datetime.datetime.strptime(str(firstyear)+'-12-31', "%Y-%m-%d")-fuelday).days
        new_var = dict_unit_year[str(firstyear)] + days_firstyear
        dict_unit_year.update({str(firstyear):new_var})
        for i in range(firstyear+1,2017):
            new_var= dict_unit_year[str(i)]+365
            dict_unit_year.update({str(i):new_var})
        if firstyear==2017:
            new_var = dict_unit_year[str(2017)] +(datetime.datetime.now()-fuelday).days
        else:
            new_var = dict_unit_year[str(2017)] +(datetime.datetime.now()-datetime.datetime.strptime(str(2016)+'-12-31', "%Y-%m-%d")).days
        dict_unit_year.update({str(2017): new_var})


def myexZero(i,o):
    if i==0:
        return o
    elif i is None:
        return o
    else:
        return i


def init(plantlist,list_RULE,column_selected):
    for plant in plantlist:
        wb=Workbook()#excel file for writting
        trash_wb=Workbook()
        trash_wb.create_sheet('unknown code',0)
        ws=[]
        ws.append(wb.create_sheet('whole', 0))
        for rule in list_RULE:
            ws.append(wb.create_sheet(rule, 0))
        wb.create_sheet('statics', 0)
        # wb.create_sheet('unknown_code', 0)
        dict_wb.update({plant:wb})
        dict_trash_wb.update({plant:trash_wb})

        dic_statics={}
        report_rules={}
        WANO_Distribution= {'根因': {'whole':{}},
                            '直接原因': {'whole':{}},
                            '系统': {'whole':{}},
                            '设备': {'whole':{}},
                            }

        for rule in list_RULE:
            report_rules.update({rule: [0, 0]})
            for factor in WANO_Distribution.items():
                factor[1].update({rule:{}})

        dic_statics.update({'WANO_Distribution':WANO_Distribution})
        dic_statics.update({'report_rules':report_rules})
        dic_statics.update({'whole':0})
        dic_statics.update({'Annual Distribution':{'whole':{}}})
        dic_statics.update({'Totol Event':0})
        SAMPLE={
            'WANO Dstribution':
                {
                    'ROOTCAUSE':
                        {
                            '4.1.1':
                                {
                                    '500':1,
                                    '600':5
                                }
                        }
                }
        }
        dict_whole_static.update({plant:dic_statics})


def myXLWR_ratio(dic_statics):
    for rule in dic_statics['report_rules'].items():
        event=rule[1][0]
        dic_statics['report_rules'].update({rule[0]:[event,float(event)/myexZero(dic_statics['whole'],1)]})
        if rule[0] in dic_statics['Annual Distribution'].keys():
            for year in dic_statics['Annual Distribution'][rule[0]].items():
                event=year[1][0]
                dic_statics['Annual Distribution'][rule[0]].update({year[0]:[event,float(event)/myexZero(dic_statics['Annual Distribution']['whole'][year[0]],1)]})


def WR_Sheet(owb,dic_statics,report_rule,report_year,data,WANO_data,trash_ws,trash_data):
    owb['whole'].append(data)
    dic_statics.update({'whole': dic_statics.get('whole') + 1})
    WR_WANO('whole', dic_statics, WANO_data, WANOmodel=wanoModel,trash_ws=trash_ws,trash_data=trash_data)

    if report_year in dic_statics['Annual Distribution']['whole'].keys():
        dic_statics['Annual Distribution']['whole'].update(
            {report_year: dic_statics['Annual Distribution']['whole'][report_year] + 1})
    else:
        dic_statics['Annual Distribution']['whole'].update({report_year: 1})

    if report_rule is not None :
        list_rule = re.split('[,，/]+',report_rule)
        for i_rule in list_rule:
            WR_Rule(i_rule, dic_statics, report_year, owb, data)
            WR_WANO(i_rule, dic_statics, WANO_data, WANOmodel=wanoModel,trash_ws=None,trash_data=None)
    else:
        WR_Rule('other', dic_statics, report_year, owb, data)
        WR_WANO('other', dic_statics, WANO_data, WANOmodel=wanoModel,trash_ws=None,trash_data=None)

def WR_Rule(rule,dic_statics,report_year,owb,data):
    new_val = dic_statics['report_rules'].get(rule)[0] + 1
    dic_statics['report_rules'].update({rule: [new_val, new_val / dic_statics['whole']]})
    if rule in dic_statics['Annual Distribution'].keys():
        if report_year in dic_statics['Annual Distribution'][rule].keys():
            new_val = dic_statics['Annual Distribution'][rule][report_year][0] + 1
            dic_statics['Annual Distribution'][rule].update(
                {report_year: [new_val, new_val / dic_statics['Annual Distribution']['whole'][report_year]]})
        else:
            dic_statics['Annual Distribution'][rule].update(
                {report_year: [1, 1 / dic_statics['Annual Distribution']['whole'][report_year]]})
    else:
        dic_statics['Annual Distribution'].update(
            {
                rule: {report_year: [1,
                                     1 / dic_statics['Annual Distribution']['whole'][report_year]
                                     ]
                    }
             }
        )
    owb[rule].append(data)


def WR_WANO(rule,dic_statics,WANO_data,WANOmodel,trash_ws,trash_data):
    unknowncodeornot = False
    for factor in WANO_data.items():  # Factor such as WANO_SYSTEM:{}
        old_factor=dic_statics['WANO_Distribution'][factor[0]][rule]
        if factor[1]:
            for i_factor in WANOmodel.code_Kind_level1(factor[1],factor[0]):
                WR_DIC_FACTOR(i_factor,old_factor)
                if i_factor=='other':
                    unknowncodeornot=True
        else:
            WR_DIC_FACTOR('other', old_factor)
            unknowncodeornot=True
    if trash_data is not None and unknowncodeornot:
            trash_ws.append(trash_data)
#     if trash_data is not None and not unknowncodeornot:
#             trash_ws.append(trash_data)

def WR_DIC_FACTOR(i_factor,old_factor):
    if str(i_factor) in old_factor.keys():
        old_value = old_factor[str(i_factor)]
        old_factor.update({str(i_factor): old_value + 1})
    else:
        old_factor.update({str(i_factor): 1})

def output_statics(workbook,statics_data):  # input statics data
    for data in sorted(statics_data['report_rules'].items(),key=lambda  item:item[0]):
        data_row=[data[0],data[1][0],data[1][1]]
        # print('going to append data for '+str(data[0]))
        workbook['statics'].append(data_row)
    workbook['statics'].append(['Totol Event',statics_data['whole']])
    list_year = [int(year) for year in statics_data['Annual Distribution']['whole'].keys()]
    for year_data in statics_data['Annual Distribution'].items():   # input year_data example year_data: ('4.1.1',{'1997':3}) tuple
        # print('going to append year_data' + str(year_data[0]))
        workbook[year_data[0]].append(['Annual Distribution'])
        workbook[year_data[0]].append(['Year','Event','Ratio'])
        if year_data[1]:  # example year_data('4.1.1',{'1997':3,'2007':6}) single tuple
            # example year_data[1]:{'1997':3,'2007':6} dict
            for i in range(min(list_year),max(list_year)):  # makeup blank year
                if str(i) not in year_data[1].keys():
                    if year_data[0]!='whole':
                        year_data[1].update({str(i):[0,0]})
                    else:
                        year_data[1].update({str(i):0})
            if year_data[0]!='whole':
                for year_data_row in sorted(year_data[1].items(),key=lambda item:item[0]):
                    workbook[year_data[0]].append([year_data_row[0],year_data_row[1][0],year_data_row[1][1]])
            else:
                for year_data_row in sorted(year_data[1].items(), key=lambda item: item[0]):
                    workbook[year_data[0]].append([year_data_row[0], year_data_row[1]])

    for wano_data in statics_data['WANO_Distribution'].items():
        for ws in workbook:
            ws.append([wano_data[0]])
            ws.append(['CODE','EVENT'])
        print('going to append wana_data' + str(wano_data[0]))
        for factor in wano_data[1].items():
            for i_factor in sorted(factor[1].items(),key=lambda item:item[0]):
                workbook[factor[0]].append([i_factor[0],i_factor[1]])


def myplot_single(filename, data_y, xticklabel, ylabel, path):
    # xticklabel.insert(0, 'blank')
    font = {'family': 'serif',
            'color': 'darkred',
            'weight': 'normal',
            'size': 26,
            }
    N=len(xticklabel)
    N_Year=5#年份横坐标显示N个
    N_Event=10#纵坐标显示N个

    ind = np.arange(0,N)  # the x locations for the groups
    width =0.5  # the width of the bars
    fig, ax = plt.subplots(figsize=(30,18), facecolor="white")
    # plt.figure(figsize=(200, 160))
    if data_y:

        ax.set_ylabel(ylabel,fontdict=font)
        ax.set_title(filename,fontsize=40,family='serif')
        ax.yaxis.grid(True, which='minor')
        ax.yaxis.grid(True, which='major')

        # plot YearDistribution Char x_label is Year
        if 'Annual Distribution' in filename:
            Skip=divmod(N,N_Year)
            xlocator=MultipleLocator(max([1,Skip[0]]))
            ind=np.array([int(x) for x in xticklabel])
            rects1 = ax.bar(ind-width/2, data_y, width,color='b')
            ax.xaxis.set_major_locator(xlocator)
            xmajorFormatter = '%d'
            ax.xaxis.set_major_formatter(FormatStrFormatter(xmajorFormatter))
            for tick in ax.xaxis.get_major_ticks():
                tick.label1.set_fontsize(20)


        # plot Rule Distribution Char x_label is rule
        else:
            rects1 = ax.bar(ind-width/2, data_y, width, color='b')
            ax.set_xticks(ind)
            ax.set_xticklabels(xticklabel,fontsize=(25-N/2))

        # if chart is Ratio, set Y format to %
        if ylabel=='Ratio':
            ymajorFormatter = '%.2f%%'
            ylocator = MultipleLocator(20)
            plt.ylim(0, 115)

        # if chart is Event, set Y format to iterger
        else:
            ymajorFormatter = '%d'
            Skip = divmod(max(data_y), N_Event)
            ylocator = MultipleLocator(max(1,divmod(Skip[0],5)[0]*5))
            plt.ylim(0, max(int(x) for x in data_y) * 1.5)

        # ax2=ax.twinx()
        # rects2 =ax2.bar(ind+width/2, data_y, width, color='R')   # this is the important function
        # ax2.set_ylim([0,  max(int(x) for x in data_y) * 1.5])
        # ax2.set_ylabel('Y values for ln(x)')
        # ax2.set_xlabel('test')
        ax.yaxis.set_major_formatter(FormatStrFormatter(ymajorFormatter))
        ax.yaxis.set_major_locator(ylocator)
        plt.xlim(min(ind) - 0.5, max(ind)  + 0.5)
        for tick in ax.yaxis.get_major_ticks():
            tick.label1.set_fontsize(25)

        ax.legend([rects1,], [ylabel,],fontsize=20)
        for rect in rects1:
            height = rect.get_height()
            plt.text(rect.get_x() + rect.get_width() / 2.,1.03 * height, ymajorFormatter % height,ha='center',fontsize=20, family='serif', style='italic', wrap=True)
    png_title = filename + '.PNG'
    fig.savefig(path + '/' + png_title)
    plt.close()


def myplot_YearDistribution(plant,data_plant,path):
    # plot Annual Distribution chart
    for report_rule in data_plant['Annual Distribution'].items():
        sorted_tuple = sorted(report_rule[1].items(), key=lambda item: item[0])
        if report_rule[0]!='whole':
            pass
            data_x=[str(x[0]) for x in sorted_tuple]
            data_y=[int(y[1][0]) for y in sorted_tuple]
            filename=plant+'_'+report_rule[0]+'_Annual Distribution_Num'
            Ylabel='Event'
            myplot_YearDistribution_single(filename,data_y,data_x,Ylabel,path)

            data_y = [float(y[1][1])*100. for y in sorted_tuple]
            filename = plant + '_' + report_rule[0] + '_Annual Distribution_Ratio'
            Ylabel = 'Ratio'
            myplot_YearDistribution_single(filename, data_y, data_x,Ylabel, path)
        else:
            data_x=[str(x[0]) for x in sorted_tuple]
            # data_y=[float(365.0)*float(y[1])/float(dict_unit_year[str(y[0])]) for y in sorted_tuple]
            data_y=[float(y[1]) for y in sorted_tuple]
            filename=plant+'_'+'whole_Annual Distribution_Num'
            Ylabel='Event'
            myplot_YearDistribution_single(filename,data_y,data_x,Ylabel,path)


def myplot_YearDistribution_single(filename,data_y,xticklabel,ylabel,path):
    font = {'family': 'serif',
            'color': 'darkred',
            'weight': 'normal',
            'size': 26,
            }
    N = len(xticklabel)
    N_Year = 5  # 年份横坐标显示N个
    N_Event = 10  # 纵坐标显示N个

    ind = np.arange(0, N)  # the x locations for the groups
    width = 0.5  # the width of the bars
    fig, ax = plt.subplots(figsize=(30, 18), facecolor="white")
    # plt.figure(figsize=(200, 160))
    if data_y:

        ax.set_ylabel(ylabel, fontdict=font)
        ax.set_title(filename, fontsize=40, family='serif')
        ax.yaxis.grid(True, which='minor')
        ax.yaxis.grid(True, which='major')

        # plot YearDistribution Char x_label is Year

        Skip = divmod(N, N_Year)
        xlocator = MultipleLocator(max([1, Skip[0]]))
        ind = np.array([int(x) for x in xticklabel])
        rects1 = ax.bar(ind - width / 2, data_y, width, color='b')
        ax.xaxis.set_major_locator(xlocator)
        xmajorFormatter = '%d'
        ax.xaxis.set_major_formatter(FormatStrFormatter(xmajorFormatter))
        for tick in ax.xaxis.get_major_ticks():
            tick.label1.set_fontsize(20)

        # if chart is Ratio, set Y format to %
        if ylabel == 'Ratio':
            ymajorFormatter = '%.2f%%'
            ylocator = MultipleLocator(20)
            plt.ylim(0, 115)

        # if chart is Event, set Y format to iterger
        else:
            ymajorFormatter = '%d'
            Skip = divmod(max(data_y), N_Event)
            ylocator = MultipleLocator(max(1, divmod(Skip[0], 5)[0] * 5))
            plt.ylim(0, max(int(x) for x in data_y) * 1.5)

        ax.yaxis.set_major_formatter(FormatStrFormatter(ymajorFormatter))
        ax.yaxis.set_major_locator(ylocator)
        plt.xlim(min(ind) - 0.5, max(ind) + 0.5)
        for tick in ax.yaxis.get_major_ticks():
            tick.label1.set_fontsize(25)

        ax.legend([rects1, ], [ylabel, ], fontsize=20)
        for rect in rects1:
            height = rect.get_height()
            plt.text(rect.get_x() + rect.get_width() / 2., 1.03 * height, ymajorFormatter % height, ha='center',
                     fontsize=20, family='serif', style='italic', wrap=True)
    png_title = filename + '.PNG'
    fig.savefig(path + '/' + png_title)
    plt.close()


def myplot_ReportRule(plant,data_plant,path):
    for x in data_plant['report_rules'].items():
        sorted_tuple = sorted(data_plant['report_rules'].items(), key=lambda item: item[0])
        data_y = [(int(y[1][0])) for y in sorted_tuple]
        Xlabel=[str(x[0]) for x in sorted_tuple]
        filename = plant + '_' + '_Report Rule Distribution_Num'
        Ylabel = 'Event'
        myplot_single(filename, data_y, Xlabel, Ylabel, path)


def myplot_ReportRule_single(filename,data_y,xticklabel,ylabel,path):
    font = {'family': 'serif',
            'color': 'darkred',
            'weight': 'normal',
            'size': 26,
            }
    N = len(xticklabel)
    N_Year = 5  # 年份横坐标显示N个
    N_Event = 10  # 纵坐标显示N个

    ind = np.arange(0, N)  # the x locations for the groups
    width = 0.5  # the width of the bars
    fig, ax = plt.subplots(figsize=(30, 18), facecolor="white")
    # plt.figure(figsize=(200, 160))
    if data_y:
        rects1 = ax.bar(ind - width / 2, data_y, width, color='b')
        ax.set_xticks(ind)
        ax.set_xticklabels(xticklabel, fontsize=(25 - N / 2))

        # if chart is Ratio, set Y format to %
        if ylabel == 'Ratio':
            ymajorFormatter = '%.2f%%'
            ylocator = MultipleLocator(20)
            plt.ylim(0, 115)

            # if chart is Event, set Y format to iterger
        else:
            ymajorFormatter = '%d'
            Skip = divmod(max(data_y), N_Event)
            ylocator = MultipleLocator(max(1, divmod(Skip[0], 5)[0] * 5))
            plt.ylim(0, max(int(x) for x in data_y) * 1.5)

        ax.yaxis.set_major_formatter(FormatStrFormatter(ymajorFormatter))
        ax.yaxis.set_major_locator(ylocator)
        plt.xlim(min(ind) - 0.5, max(ind) + 0.5)
        for tick in ax.yaxis.get_major_ticks():
            tick.label1.set_fontsize(25)

        ax.legend([rects1, ], [ylabel, ], fontsize=20)
        for rect in rects1:
            height = rect.get_height()
            plt.text(rect.get_x() + rect.get_width() / 2., 1.03 * height, ymajorFormatter % height, ha='center',
                     fontsize=20, family='serif', style='italic', wrap=True)


    png_title = filename + '.PNG'
    fig.savefig(path + '/' + png_title)
    plt.close()


def myplot_firstyear(path):
    font = {'family': 'serif',
            'color': 'darkred',
            'weight': 'normal',
            'size': 26,
            }
    x_data=[]
    y_data=[]
    y_ticks=[]
    for item in PlantInfo.items():
        x_data.append(datetime.datetime.strptime((item[1][3]), "%Y-%m-%d"))
        y_data.append(int(dict_whole_static[item[0]]['whole']))
        y_ticks.append(str(item[0]))
    fig, ax = plt.subplots(figsize=(30, 18), facecolor="white")
    sact1=ax.scatter(x_data, y_data, marker='s', s=150)
    ax.set_ylabel('Event', fontdict=font)
    ax.set_xlabel('FFD',fontdict=font)
    ax.set_title('Events Occured in the First Year After FFD ', fontsize=40, family='serif')
    ax.yaxis.grid(True, which='minor')
    ax.yaxis.grid(True, which='major')
    for tick in ax.yaxis.get_major_ticks():
        tick.label1.set_fontsize(25)

    for tick in ax.xaxis.get_major_ticks():
        tick.label1.set_fontsize(20)

    Skip = divmod(max(y_data), 5)
    ylocator = MultipleLocator(max(1, divmod(Skip[0], 5)[0] * 5))
    ax.yaxis.set_major_locator(ylocator)
    for i in range(0,len(x_data)):
        text(x_data[i], y_data[i]-0.1 ,y_ticks[i], fontsize=15)
    plt.ylim(0, max(y_data)*1.3)
    ax.yaxis.set_major_formatter(FormatStrFormatter('%d'))
    png_title = 'firstyear' + '.PNG'
    fig.savefig(path + '/' + png_title)

    plt.close()


def myplot_WANO(plant, data_plant, path):
    for data in data_plant.items():
        for item in data[1].items():
            if item[1]:
                myplot_WANO_single(plant+'_'+str(item[0])+'_'+str(data[0]),data[0],item[1],path)


def myplot_WANO_single(title, area,data_plant, path):
    font = {'family': 'serif',
            'color': 'darkred',
            'weight': 'normal',
            'size': 40,
            }
    data_sorted=sorted(data_plant.items(),key=lambda item:int(item[1]),reverse=True)
    value=[int(data[1]) for data in data_sorted]
    label=[wanoModel.code[area][str(data[0])] for data in data_sorted]

    fig, ax = plt.subplots(figsize=(30, 12), facecolor="white")
    exp=[0.1]
    for i in range(0,len(value)-1):
        exp.append(0)
    patches, l_text, p_text=ax.pie(value,labels=label,explode=exp,autopct = '%3.1f%%',shadow = False,startangle = 90,pctdistance = 0.6)
    ax.axis('equal')
    ax.legend()
    for t in l_text:
        t.set_size = (40)
    for t in p_text:
        t.set_size = (40)
    png_title = title + '.PNG'
    fig.savefig(path + '/' + png_title)
    plt.close()


def myplot_plant(plant,data_plant,path):
    # plot Annual Distribution chart
    myplot_YearDistribution(plant,data_plant,path)
    # Plot report rule Distribution. just 1 chart
    myplot_ReportRule(plant,data_plant,path)


def myXLWR_n(plant, input_row,column_selected,dict_trash_wb=dict_trash_wb):  # write single row to the purpose wb and update the dic_statics
    data = []
    nr = len(input_row)
    # event_begin=datetime.datetime.strptime(input_row[Tiltle2index('EVENT_BEGIN')].value[0:10],"%Y-%m-%d")
    # firstyear=datetime.datetime.strptime(PlantInfo[plant][3],"%Y-%m-%d")
    if True:#'4.1.4' in input_row[Tiltle2index('REPORT_RULE')].value:#event_begin<=firstyear+datetime.timedelta(days=365):#period screening
    # if event_begin <=datetime.datetime.strptime(PlantInfo[plant][2],"%Y-%m-%d"):  # period screening
        if True:
            if str(input_row[Tiltle2index('flag')].value)!='1': #and event_begin <=datetime.datetime.strptime(PlantInfo[plant][2],"%Y-%m-%d"):#period screening
                for j in range(0, nr):
                    data.append(input_row[j].value)
                report_rule = input_row[Tiltle2index('报告准则')].value
                report_year = str(input_row[Tiltle2index('年度')].value)
                WANO_data = {'根因': str(input_row[Tiltle2index('根因')].value),
                             '直接原因': str(input_row[Tiltle2index('直接原因')].value),
                             '系统': str(input_row[Tiltle2index('系统')].value),
                             '设备': str(input_row[Tiltle2index('设备')].value)
                             }
                WR_Sheet(dict_wb['ALLPlant'], dict_whole_static['ALLPlant'], report_rule, report_year, data, WANO_data,trash_ws=dict_trash_wb['ALLPlant']['unknown code'],trash_data=data)
                WR_Sheet(dict_wb[plant], dict_whole_static[plant], report_rule, report_year, data, WANO_data,trash_ws=dict_trash_wb[plant]['unknown code'],trash_data=data)
                new_tot=dict_whole_static[plant]['Totol Event']+1
                dict_whole_static[plant].update({'Totol Event':new_tot})


test=False
if not test:
    init(list_Plant,list_RULE,columsPurpose)
    caculate_unityear()
    for tuple_wbi in dict_wb.items():#input tilte row
        for wsi in tuple_wbi[1]:
            if wsi.title!='statics':
                wsi.append(listTitle)

    for i in range(1,rs_rows_len):
        row=list(list(rs.rows)[i])
        plant=row[Tiltle2index('机组')].value
        myXLWR_n(plant, row,columsPurpose,dict_trash_wb)
        print('row '+str(i)+' is completed')

    pass
    conn = '/after 1 year in FFD'
    for dict_wbi in dict_wb.items():#input statics and save
        plant = dict_wbi[0]
        if True:
            myXLWR_ratio(dict_whole_static[plant])
            output_statics(dict_wbi[1],dict_whole_static[plant])
            if not os.path.exists(filepath+'/'+plant+conn):
                os.makedirs(filepath+'/'+plant+conn)
            myplot_WANO(plant,dict_whole_static[plant]['WANO_Distribution'],filepath+'/'+plant+conn)
            myplot_plant(plant,dict_whole_static[plant],filepath+'/'+plant+conn)
            dict_wbi[1].save(filepath+'/'+plant+conn+'/'+plant+'.xlsx')
            dict_trash_wb[plant].save(filepath+'/'+plant+conn+'/'+plant+'unknown_code.xlsx')
            print(plant+' data has been saved')

    myplot_firstyear(filepath + '/' + 'ALLPlant' + conn)
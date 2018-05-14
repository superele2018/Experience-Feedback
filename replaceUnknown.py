from openpyxl import Workbook
from openpyxl import load_workbook
import time
from openpyxl.chart import BarChart, Series, Reference
from pylab import *
from matplotlib.ticker import MultipleLocator
from matplotlib.ticker import FormatStrFormatter
import os
import WanoCode
import datetime


wb_new=Workbook()
ws_new=wb_new.worksheets.create_sheet('unknown code',0)
wb_old=load_workbook('input.xlsx')
ws_old=wb_old.worksheets[0]
wb_replace=load_workbook('replace.xlsx')
ws_replace=wb_replace.worksheets[0]


def getdatafromsheet(sheet):
    list_row=[]
    for row in list(sheet.rows):
        list_data=[]
        for cell in list(row):
            list_data.append(cell.value)
        list_row.append(list_data)
    return list_row


data_old=getdatafromsheet(ws_old)
data_new=getdatafromsheet(ws_new)
data_replace=getdatafromsheet(ws_replace)

for i in range(0,len(data_old)):
    if data_old





